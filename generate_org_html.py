#!/usr/bin/env python3
"""
Org Chart HTML Generator — Master Excel Edition

Reads a single orgchart_master_data.xlsx (3 sheets: People, Scrum Teams, Teams Hierarchy)
and produces the same org_drilldown.html + org_drilldown_redacted.html as the original
multi-source generator.

No fuzzy matching, no name aliases, no multi-source merging — all data is pre-normalized
in the master Excel.
"""

import json
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# Import shared functions
from org_html_shared import (
    normalize_name,
    slugify,
    title_seniority_score,
    redact_data,
    verify_redaction,
    generate_html,
)

# ─── Configuration ───────────────────────────────────────────────────────────

MASTER_FILE = Path(__file__).parent / "data" / "orgchart_master_data.xlsx"
OUTPUT_FILE = Path(__file__).parent / "org_drilldown.html"
REDACTED_FILE = Path(__file__).parent / "org_drilldown_redacted.html"

JAYESH_NAME = "Jayesh Sahasi"
JAYESH_TITLE = "Executive VP, Products and CTO"

# Canonical org order (matches original generator iteration order)
ORG_ORDER = ["Product-Design", "Full QA Org", "Full Dev Org", "Salesforce", "TPM"]

# All discipline groups that the HTML template expects (even if empty)
ALL_DISCIPLINES = ["Dev", "QA", "Product", "TPM", "Salesforce", "Other"]


# ─── Step 1: Parse People sheet ─────────────────────────────────────────────

def parse_people(filepath):
    """Read the People sheet. Returns {norm_name: person_dict}."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["People"]

    # Read header row to find column indices
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    people = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col["Name"]]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        norm = normalize_name(name)
        if not norm:
            continue

        scrum_raw = row[col.get("Scrum Teams", -1)] if "Scrum Teams" in col else ""
        scrum_teams = [t.strip() for t in (scrum_raw or "").split(";") if t.strip()]

        people[norm] = {
            "name": name,
            "title": (row[col["Title"]] or "").strip() if row[col["Title"]] else "",
            "employment": (row[col["Employment"]] or "").strip() if row[col["Employment"]] else "",
            "org": (row[col["Org"]] or "").strip() if row[col["Org"]] else "",
            "reportsTo": (row[col["Reports To"]] or "").strip() if row[col["Reports To"]] else "",
            "scrumTeams": scrum_teams,
            "location": (row[col.get("Location", -1)] or "").strip() if col.get("Location") is not None and row[col["Location"]] else "",
            "talentBand": (row[col.get("Talent Band", -1)] or "").strip() if col.get("Talent Band") is not None and row[col["Talent Band"]] else "",
            "talentCategory": (row[col.get("Cvent Talent Category", -1)] or "").strip() if col.get("Cvent Talent Category") is not None and row[col["Cvent Talent Category"]] else "",
            "rationale": (row[col.get("Rationale", -1)] or "").strip() if col.get("Rationale") is not None and row[col["Rationale"]] else "",
        }

    wb.close()
    return people


# ─── Step 2: Parse Scrum Teams sheet ────────────────────────────────────────

def parse_scrum_teams(filepath):
    """Read the Scrum Teams sheet.
    Returns (teams_dict, meta_dict) where:
      teams_dict = {team_name: {discipline: [{name, isLead}]}}
      meta_dict  = {team_name: {"scrumMaster": str, "productOwner": str}}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Scrum Teams"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    teams = defaultdict(lambda: defaultdict(list))
    meta = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        team = row[col["Team Name"]]
        member = row[col["Member Name"]]
        if not team or not member:
            continue
        team = str(team).strip()
        member = str(member).strip()
        discipline = (row[col["Discipline"]] or "").strip() if row[col["Discipline"]] else "Other"
        is_lead_raw = (row[col["Is Lead"]] or "").strip() if row[col["Is Lead"]] else "No"
        is_lead = is_lead_raw.lower() in ("yes", "true", "1")

        teams[team][discipline].append({
            "name": member,
            "isLead": is_lead,
        })

        # Capture team-level SM/PO metadata (first non-empty value wins)
        if team not in meta:
            po_col = col.get("Product Owner")
            sm_col = col.get("Scrum Master")
            po = (str(row[po_col]).strip() if po_col is not None and row[po_col] else "")
            sm = (str(row[sm_col]).strip() if sm_col is not None and row[sm_col] else "")
            meta[team] = {"productOwner": po, "scrumMaster": sm}

    wb.close()
    return dict(teams), meta


# ─── Step 3: Build org datasets ─────────────────────────────────────────────

def build_org_datasets(people):
    """Build the orgs dict from flat People data.

    Returns {org_name: {top, nodes, children}} matching the DATA JSON contract.
    """
    jayesh_norm = normalize_name(JAYESH_NAME)

    # Create node IDs and build lookup
    norm_to_id = {}
    norm_to_person = {}
    for norm, person in people.items():
        nid = slugify(person["name"])
        # Handle duplicate slugs
        if nid in {v for v in norm_to_id.values()}:
            nid = nid + "-2"
        norm_to_id[norm] = nid
        norm_to_person[norm] = person

    # Group people by org
    orgs_people = defaultdict(list)
    for norm, person in people.items():
        org = person["org"]
        if org:
            orgs_people[org].append(norm)

    # Build parent→children from Reports To
    children_all = defaultdict(list)  # parent_norm → [child_norms]
    for norm, person in people.items():
        mgr_name = person["reportsTo"]
        if mgr_name:
            mgr_norm = normalize_name(mgr_name)
            if mgr_norm in norm_to_id:
                children_all[mgr_norm].append(norm)

    jayesh_id = norm_to_id.get(jayesh_norm, "jayesh-sahasi")

    # Build per-org datasets in canonical order
    org_datasets = {}
    for org_name in ORG_ORDER:
        member_norms = orgs_people.get(org_name)
        if not member_norms:
            continue
        # Collect all nodes in this org (+ Jayesh as top)
        org_norms = set(member_norms)
        org_norms.add(jayesh_norm)

        nodes = {}
        children = {}

        for norm in org_norms:
            person = people.get(norm)
            if not person:
                continue
            nid = norm_to_id[norm]
            # Jayesh's org field should match the current org dataset
            node_org = org_name if norm == jayesh_norm else person["org"]
            nodes[nid] = {
                "id": nid,
                "name": person["name"],
                "title": person["title"],
                "employment": person["employment"],
                "scrumTeams": person["scrumTeams"],
                "placeholder": False,
                "org": node_org,
                "talentBand": person["talentBand"],
                "talentCategory": person["talentCategory"],
                "rationale": person["rationale"],
            }

        # Build children map for this org (only nodes within this org)
        for norm in org_norms:
            nid = norm_to_id.get(norm)
            if not nid or nid not in nodes:
                continue
            kids = []
            for child_norm in children_all.get(norm, []):
                child_id = norm_to_id.get(child_norm)
                if child_id and child_id in nodes:
                    kids.append(child_id)
            if kids:
                # Sort children alphabetically by name
                kids.sort(key=lambda cid: nodes[cid]["name"].lower())
                children[nid] = kids

        org_datasets[org_name] = {
            "top": jayesh_id,
            "nodes": nodes,
            "children": children,
        }

    return org_datasets


# ─── Step 4: Build scrum data ───────────────────────────────────────────────

def build_scrum_data(people, scrum_sheet, org_datasets):
    """Build the scrum dict from the Scrum Teams sheet + People data.

    Returns {team_name: {discipline: [member_objs]}} matching the DATA JSON contract.
    """
    # Build name → node ID lookup across all orgs
    name_to_id = {}
    for org_name, ds in org_datasets.items():
        for nid, node in ds["nodes"].items():
            norm = normalize_name(node["name"])
            if norm not in name_to_id:
                name_to_id[norm] = nid

    # Build name → org lookup
    name_to_org = {}
    for norm, person in people.items():
        name_to_org[norm] = person["org"]

    scrum_teams = {}
    for team_name, disciplines in scrum_sheet.items():
        # Initialize all discipline groups (HTML template expects them all)
        grouped = {d: [] for d in ALL_DISCIPLINES}
        for discipline, members in disciplines.items():
            member_objs = []
            for m in members:
                norm = normalize_name(m["name"])
                person = people.get(norm, {})
                member_objs.append({
                    "name": m["name"],
                    "id": name_to_id.get(norm, slugify(m["name"])),
                    "org": person.get("org", name_to_org.get(norm, "")),
                    "title": person.get("title", ""),
                    "employment": person.get("employment", ""),
                    "isLead": m["isLead"],
                    "talentBand": person.get("talentBand", ""),
                    "talentCategory": person.get("talentCategory", ""),
                    "rationale": person.get("rationale", ""),
                })
            # Sort: leads first, then by seniority, then alphabetically
            member_objs.sort(key=lambda x: (
                0 if x["isLead"] else 1,
                -title_seniority_score(x["title"]),
                x["name"].lower(),
            ))
            grouped[discipline] = member_objs
        scrum_teams[team_name] = grouped

    return scrum_teams


# ─── Step 5: Build home DRs ────────────────────────────────────────────────

def build_home_drs(org_datasets):
    """Find Jayesh's direct children across all orgs, deduplicated."""
    home_drs = []
    seen = set()

    for org_name, ds in org_datasets.items():
        top_id = ds["top"]
        top_children = ds["children"].get(top_id, [])
        for cid in top_children:
            node = ds["nodes"].get(cid)
            if not node:
                continue
            norm = normalize_name(node["name"])
            if norm in seen:
                continue
            seen.add(norm)
            home_drs.append({
                "name": node["name"],
                "title": node.get("title", ""),
                "employment": node.get("employment", ""),
                "org": org_name,
                "nodeId": cid,
                "talentBand": node.get("talentBand", ""),
                "talentCategory": node.get("talentCategory", ""),
                "rationale": node.get("rationale", ""),
            })

    return home_drs


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Org Chart Generator (from Master Excel)")
    print("=" * 60)

    if not MASTER_FILE.exists():
        print(f"\nERROR: Master file not found: {MASTER_FILE}")
        print("Run generate_org_html_legacy.py first to create it, or provide the master Excel.")
        sys.exit(1)

    # Step 1: Parse People
    print("\n[1] Parsing People sheet...")
    people = parse_people(MASTER_FILE)
    print(f"  {len(people)} people loaded")

    # Step 2: Parse Scrum Teams
    print("\n[2] Parsing Scrum Teams sheet...")
    scrum_sheet, scrum_meta = parse_scrum_teams(MASTER_FILE)
    print(f"  {len(scrum_sheet)} teams loaded")
    meta_with_po = sum(1 for m in scrum_meta.values() if m.get("productOwner"))
    meta_with_sm = sum(1 for m in scrum_meta.values() if m.get("scrumMaster"))
    print(f"  {meta_with_po} teams with Product Owner, {meta_with_sm} with Scrum Master")

    # Step 3: Build org datasets
    print("\n[3] Building org datasets...")
    org_datasets = build_org_datasets(people)
    for org_name, ds in org_datasets.items():
        print(f"  {org_name}: {len(ds['nodes'])} nodes")

    # Step 4: Build scrum data
    print("\n[4] Building scrum team data...")
    scrum_teams = build_scrum_data(people, scrum_sheet, org_datasets)
    print(f"  {len(scrum_teams)} scrum teams")
    for team_name in sorted(scrum_teams.keys()):
        groups = scrum_teams[team_name]
        total = sum(len(g) for g in groups.values())
        print(f"    {team_name}: {total} members")

    # Step 5: Build home DRs
    print("\n[5] Building home DRs...")
    home_drs = build_home_drs(org_datasets)
    print(f"  {len(home_drs)} direct reports")

    # Step 6: Assemble DATA
    data = {
        "orgs": {
            org_name: {
                "top": ds["top"],
                "nodes": ds["nodes"],
                "children": ds["children"],
            }
            for org_name, ds in org_datasets.items()
        },
        "scrum": scrum_teams,
        "scrumMeta": scrum_meta,
        "homeDrs": home_drs,
        "missing": {},
    }

    # Step 7: Generate named HTML
    print("\n[6] Generating HTML files...")
    html_full = generate_html(data, redacted=False)
    OUTPUT_FILE.write_text(html_full, encoding="utf-8")
    print(f"  Written: {OUTPUT_FILE} ({len(html_full):,} bytes)")

    # Step 8: Generate redacted HTML
    all_names = set()
    for ds in org_datasets.values():
        for nid, node in ds["nodes"].items():
            if not node.get("placeholder"):
                all_names.add(node["name"])

    redacted_data = redact_data(data, all_names)
    html_redacted = generate_html(redacted_data, redacted=True)

    leaked = verify_redaction(html_redacted, all_names)
    if leaked:
        print(f"\n  [WARN] Redaction verification: {len(leaked)} names may have leaked:")
        for name in leaked[:10]:
            print(f"    - {name}")
    else:
        print("  Redaction verification: PASSED (no names found in HTML)")

    REDACTED_FILE.write_text(html_redacted, encoding="utf-8")
    print(f"  Written: {REDACTED_FILE} ({len(html_redacted):,} bytes)")

    # Summary
    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    for org_name in ["Product-Design", "Full QA Org", "Full Dev Org", "Salesforce", "TPM"]:
        if org_name in org_datasets:
            ds = org_datasets[org_name]
            print(f"  {org_name}: {len(ds['nodes'])} nodes")
    print(f"\n  Scrum teams: {len(scrum_teams)}")
    print(f"\n  Output files:")
    print(f"    {OUTPUT_FILE}")
    print(f"    {REDACTED_FILE}")
    print("\nDone!")


if __name__ == "__main__":
    main()
