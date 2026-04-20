"""Retry unresolved names: pattern-guess emails, multi-token combos, personal contacts."""
import re
import openpyxl
import win32com.client

XLSX = "data/orgchart_master_data.xlsx"


def strip_prefix(n):
    return re.sub(r"^[Cc]-\s*", "", (n or "").strip())


def is_contractor(raw):
    return bool(re.match(r"^[Cc]-", (raw or "").strip()))


def extract_email(entry):
    try:
        xu = entry.GetExchangeUser()
        if xu and xu.PrimarySmtpAddress:
            return xu.PrimarySmtpAddress
    except Exception:
        pass
    try:
        pa = entry.GetPropertyAccessor()
        v = pa.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
        if v:
            return v
    except Exception:
        pass
    return None


def try_resolve(ns, name):
    r = ns.CreateRecipient(name)
    r.Resolve()
    if r.Resolved:
        em = extract_email(r.AddressEntry)
        if em:
            return em
    return None


def verify_email(ns, addr):
    """Resolve an SMTP address; if it resolves to a known user, return canonical email."""
    try:
        r = ns.CreateRecipient(addr)
        r.Resolve()
        if r.Resolved:
            em = extract_email(r.AddressEntry)
            if em:
                return em
    except Exception:
        pass
    return None


def pattern_candidates(raw):
    """Generate plausible email local-parts for on24.com."""
    clean = strip_prefix(raw)
    parts = clean.split()
    if len(parts) < 2:
        return []
    first, last = parts[0].lower(), parts[-1].lower()
    prefix = "c-" if is_contractor(raw) else ""
    cands = []
    # Standard patterns
    cands.append(f"{prefix}{first}.{last}@on24.com")
    cands.append(f"{prefix}{first}{last}@on24.com")
    cands.append(f"{prefix}{first[0]}{last}@on24.com")
    cands.append(f"{prefix}{first}@on24.com")
    # 3-token names: try middle as last, full middle+last concatenated
    if len(parts) >= 3:
        mid = parts[1].lower()
        cands.append(f"{prefix}{first}.{mid}@on24.com")
        cands.append(f"{prefix}{first}.{mid}{last}@on24.com")
        cands.append(f"{prefix}{first}{last[0]}@on24.com")
    # Without c- prefix (in case they switched to employee)
    if prefix:
        cands.append(f"{first}.{last}@on24.com")
    return cands


def search_contacts(ns, first, last):
    """Scan personal Contacts folders for a match."""
    first_l = first.lower()
    last_l = last.lower()
    matches = []
    for i in range(1, ns.AddressLists.Count + 1):
        al = ns.AddressLists.Item(i)
        if "Contact" not in al.Name:
            continue
        try:
            entries = al.AddressEntries
            for j in range(1, entries.Count + 1):
                try:
                    e = entries.Item(j)
                    nm = (e.Name or "").lower()
                    if last_l in nm and first_l[:3] in nm:
                        em = extract_email(e)
                        if em:
                            matches.append((e.Name, em))
                except Exception:
                    pass
        except Exception:
            pass
    return matches


def main():
    app = win32com.client.Dispatch("Outlook.Application")
    ns = app.GetNamespace("MAPI")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["People"]
    h = [c.value for c in ws[1]]
    ni = h.index("Name") + 1
    ei = h.index("Email") + 1

    # Find unresolved
    unresolved = []
    for ridx in range(2, ws.max_row + 1):
        nm = ws.cell(row=ridx, column=ni).value
        em = ws.cell(row=ridx, column=ei).value
        if nm and not em:
            unresolved.append((ridx, str(nm).strip()))

    print(f"Retrying {len(unresolved)} unresolved names...\n")
    fixed = 0
    still = []

    for ridx, raw in unresolved:
        clean = strip_prefix(raw)
        parts = clean.split()
        email = None
        via = None

        # Strategy 1: multi-token combos (drop middle, swap order)
        if len(parts) >= 3:
            combos = [
                f"{parts[0]} {parts[1]}",       # first + middle
                f"{parts[0]} {parts[-1]}",      # first + last
                f"{parts[-1]} {parts[0]}",      # reversed
                " ".join(parts[:-1]),           # drop last
                " ".join(parts[1:]),            # drop first
            ]
            for c in combos:
                email = try_resolve(ns, c)
                if email:
                    via = f"combo:{c}"
                    break

        # Strategy 2: pattern guess + verify
        if not email:
            for cand in pattern_candidates(raw):
                em = verify_email(ns, cand)
                if em:
                    email = em
                    via = f"pattern:{cand}"
                    break

        # Strategy 3: personal contacts fuzzy
        if not email and len(parts) >= 2:
            matches = search_contacts(ns, parts[0], parts[-1])
            if len(matches) == 1:
                email = matches[0][1]
                via = f"contacts:{matches[0][0]}"
            elif len(matches) > 1:
                # pick exchange-style on24.com if any
                on24 = [m for m in matches if "@on24.com" in m[1].lower()]
                if len(on24) == 1:
                    email = on24[0][1]
                    via = f"contacts-on24:{on24[0][0]}"

        if email:
            ws.cell(row=ridx, column=ei, value=email)
            fixed += 1
            print(f"  OK  {raw:35s} -> {email}  [{via}]")
        else:
            still.append(raw)
            print(f"  --  {raw}")

    wb.save(XLSX)
    print(f"\nFixed {fixed}/{len(unresolved)}. Still unresolved: {len(still)}")
    for n in still:
        print(f"  - {n}")


if __name__ == "__main__":
    main()
