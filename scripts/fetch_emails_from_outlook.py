"""One-off: resolve People sheet names against Outlook GAL, write Email column."""
import re
import sys
import openpyxl
import win32com.client

XLSX = "data/orgchart_master_data.xlsx"

NICKNAMES = {
    "stephen": ["steve"], "steve": ["stephen"],
    "daniel": ["dan"], "dan": ["daniel"],
    "michael": ["mike"], "mike": ["michael"],
    "benjamin": ["ben"], "ben": ["benjamin"],
    "william": ["will", "bill"], "will": ["william"], "bill": ["william"],
    "robert": ["rob", "bob"], "rob": ["robert"], "bob": ["robert"],
    "richard": ["rick", "rich"], "rick": ["richard"], "rich": ["richard"],
    "james": ["jim", "jamie"], "jim": ["james"], "jamie": ["james"],
    "thomas": ["tom"], "tom": ["thomas"],
    "christopher": ["chris"], "chris": ["christopher"],
    "matthew": ["matt"], "matt": ["matthew"],
    "andrew": ["andy", "drew"], "andy": ["andrew"],
    "anthony": ["tony"], "tony": ["anthony"],
    "jonathan": ["jon", "john"], "jon": ["jonathan"],
    "joseph": ["joe"], "joe": ["joseph"],
    "nicholas": ["nick"], "nick": ["nicholas"],
    "alexander": ["alex"], "alex": ["alexander"],
    "samuel": ["sam"], "sam": ["samuel"],
    "elizabeth": ["liz", "beth"], "liz": ["elizabeth"],
    "jennifer": ["jen", "jenny"], "jen": ["jennifer"], "jenny": ["jennifer"],
    "katherine": ["kate", "kathy"], "kate": ["katherine"],
    "kenneth": ["ken"], "ken": ["kenneth"],
    "ronald": ["ron"], "ron": ["ronald"],
    "patrick": ["pat"], "pat": ["patrick"],
    "gregory": ["greg"], "greg": ["gregory"],
    "timothy": ["tim"], "tim": ["timothy"],
    "theodore": ["ted"], "ted": ["theodore"],
}


def strip_prefix(name):
    return re.sub(r"^[Cc]-\s*", "", (name or "").strip())


def extract_email(entry):
    try:
        xu = entry.GetExchangeUser()
        if xu and xu.PrimarySmtpAddress:
            return xu.PrimarySmtpAddress
    except Exception:
        pass
    try:
        pa = entry.GetPropertyAccessor()
        v = pa.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
        if v:
            return v
    except Exception:
        pass
    return None


def resolve(ns, name):
    r = ns.CreateRecipient(name)
    r.Resolve()
    if r.Resolved:
        em = extract_email(r.AddressEntry)
        if em:
            return em, "resolved"
    return None, None


def gal_search_last(gal_entries_cache, first, last):
    """Search cached GAL entries by last name; require first-initial to match."""
    last_l = last.lower()
    first_l = (first or "").lower()
    first_initial = first_l[:1]
    matches = []
    for name, entry in gal_entries_cache:
        nl = (name or "").lower()
        # Match "First Last" or "Last, First"
        if nl.endswith(" " + last_l) or nl.startswith(last_l + ","):
            # Extract the "first" token from GAL entry
            if "," in nl:
                gal_first = nl.split(",", 1)[1].strip().split()[0] if nl.split(",", 1)[1].strip() else ""
            else:
                gal_first = nl.split()[0] if nl.split() else ""
            # strip c- prefix from GAL first token
            gal_first = re.sub(r"^c-\s*", "", gal_first)
            if gal_first and first_initial and gal_first[0] == first_initial:
                matches.append((entry, gal_first))
    if len(matches) == 1:
        em = extract_email(matches[0][0])
        if em:
            return em, 1
    # If multiple but exactly one has exact first-name match, prefer it
    exact = [m for m in matches if m[1] == first_l]
    if len(exact) == 1:
        em = extract_email(exact[0][0])
        if em:
            return em, 1
    return None, len(matches)


def build_gal_cache(ns):
    gal = None
    for i in range(1, ns.AddressLists.Count + 1):
        al = ns.AddressLists.Item(i)
        if "Global Address List" in al.Name:
            gal = al
            break
    if not gal:
        return []
    entries = gal.AddressEntries
    cache = []
    total = entries.Count
    print(f"  indexing {total} GAL entries...", flush=True)
    for i in range(1, total + 1):
        try:
            e = entries.Item(i)
            cache.append((e.Name or "", e))
        except Exception:
            pass
    return cache


def main():
    print("Connecting to Outlook...")
    app = win32com.client.Dispatch("Outlook.Application")
    ns = app.GetNamespace("MAPI")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["People"]
    headers = [c.value for c in ws[1]]

    if "Email" not in headers:
        email_col = len(headers) + 1
        ws.cell(row=1, column=email_col, value="Email")
        headers.append("Email")
        print(f"Added Email column at position {email_col}")
    else:
        email_col = headers.index("Email") + 1

    name_col = headers.index("Name") + 1

    # Collect names
    rows = []
    for ridx in range(2, ws.max_row + 1):
        nm = ws.cell(row=ridx, column=name_col).value
        if nm and str(nm).strip():
            rows.append((ridx, str(nm).strip()))
    print(f"Resolving {len(rows)} names...")

    gal_cache = None
    stats = {"resolved": 0, "nickname": 0, "lastname": 0, "unresolved": 0, "ambiguous": 0}
    unresolved = []

    for ridx, raw in rows:
        clean = strip_prefix(raw)
        # Keep original (with C- prefix) as first try — contractors are in GAL with C-
        email = None
        status = None
        for attempt in [raw, clean]:
            email, status = resolve(ns, attempt)
            if email:
                break

        if not email:
            # Nickname substitution on first name
            parts = clean.split()
            if len(parts) >= 2:
                first = parts[0].lower()
                last = " ".join(parts[1:])
                for alt in NICKNAMES.get(first, []):
                    cand = f"{alt.title()} {last}"
                    email, status = resolve(ns, cand)
                    if email:
                        status = "nickname"
                        break

        if not email:
            # Last-name GAL search
            parts = clean.split()
            if len(parts) >= 2:
                if gal_cache is None:
                    gal_cache = build_gal_cache(ns)
                last = parts[-1]
                first = parts[0]
                em, n_match = gal_search_last(gal_cache, first, last)
                if em:
                    email, status = em, "lastname"
                elif n_match > 1:
                    status = "ambiguous"

        if email:
            ws.cell(row=ridx, column=email_col, value=email)
            if status in stats:
                stats[status] += 1
            else:
                stats["resolved"] += 1
            print(f"  OK  [{status:9s}] {raw:35s} -> {email}")
        else:
            if status == "ambiguous":
                stats["ambiguous"] += 1
                print(f"  ??  [ambiguous] {raw}")
            else:
                stats["unresolved"] += 1
                print(f"  --  [no-match ] {raw}")
            unresolved.append(raw)

    wb.save(XLSX)
    print("\n=== Summary ===")
    for k, v in stats.items():
        print(f"  {k:12s}: {v}")
    print(f"  total       : {len(rows)}")
    if unresolved:
        print("\nUnresolved names:")
        for n in unresolved:
            print(f"  - {n}")


if __name__ == "__main__":
    main()
