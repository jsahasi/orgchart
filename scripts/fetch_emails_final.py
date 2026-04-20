"""Final-pass: resolve unresolved names by scanning ALL address lists (not just offline GAL)."""
import re
import openpyxl
import win32com.client

XLSX = "data/orgchart_master_data.xlsx"


def strip_prefix(n):
    return re.sub(r"^[Cc]-\s*", "", (n or "").strip())


def extract_email(entry):
    try:
        xu = entry.GetExchangeUser()
        if xu and xu.PrimarySmtpAddress:
            return xu.PrimarySmtpAddress
    except Exception:
        pass
    try:
        pa = entry.GetPropertyAccessor()
        v = pa.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
        if v:
            return v
    except Exception:
        pass
    return None


def build_combined_index(ns):
    """Flatten every address list into one (normalized_name → entry) map."""
    index = {}  # lower-cased full name -> entry
    for i in range(1, ns.AddressLists.Count + 1):
        al = ns.AddressLists.Item(i)
        try:
            entries = al.AddressEntries
            total = entries.Count
        except Exception:
            continue
        for j in range(1, total + 1):
            try:
                e = entries.Item(j)
                nm = (e.Name or "").strip()
                if not nm:
                    continue
                key = nm.lower()
                if key not in index:
                    index[key] = e
            except Exception:
                pass
    print(f"  combined index: {len(index)} unique names across all lists")
    return index


def norm(s):
    return re.sub(r"[^a-z ]", "", (s or "").lower()).strip()


def find(index, raw):
    """Try: exact (with/without prefix), case-insensitive variants."""
    clean = strip_prefix(raw)
    candidates = [
        raw.lower(),
        clean.lower(),
        f"c-{clean}".lower(),
        f"C-{clean}".lower(),
    ]
    # Also "Last, First"
    parts = clean.split()
    if len(parts) >= 2:
        candidates.append(f"{parts[-1]}, {parts[0]}".lower())
        candidates.append(f"c-{parts[-1]}, {parts[0]}".lower())

    for c in candidates:
        if c in index:
            em = extract_email(index[c])
            if em:
                return em, "exact"

    # Loose match: normalized full-name equality
    target = norm(clean)
    for k, e in index.items():
        if norm(strip_prefix(k)) == target:
            em = extract_email(e)
            if em:
                return em, "normalized"
    return None, None


def main():
    app = win32com.client.Dispatch("Outlook.Application")
    ns = app.GetNamespace("MAPI")

    print("Building combined index from ALL address lists...")
    index = build_combined_index(ns)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["People"]
    h = [c.value for c in ws[1]]
    ni = h.index("Name") + 1
    ei = h.index("Email") + 1

    unresolved = []
    for ridx in range(2, ws.max_row + 1):
        nm = ws.cell(row=ridx, column=ni).value
        em = ws.cell(row=ridx, column=ei).value
        if nm and not em:
            unresolved.append((ridx, str(nm).strip()))

    print(f"\nRetrying {len(unresolved)} unresolved...\n")
    fixed = 0
    still = []
    for ridx, raw in unresolved:
        email, via = find(index, raw)
        if email:
            ws.cell(row=ridx, column=ei, value=email)
            fixed += 1
            print(f"  OK  [{via:10s}] {raw:35s} -> {email}")
        else:
            still.append(raw)
            print(f"  --              {raw}")

    wb.save(XLSX)
    print(f"\nFixed {fixed}/{len(unresolved)}. Still unresolved: {len(still)}")
    for n in still:
        print(f"  - {n}")


if __name__ == "__main__":
    main()
