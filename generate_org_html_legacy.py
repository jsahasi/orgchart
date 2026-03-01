#!/usr/bin/env python3
"""
Org Chart HTML Generator (Legacy)
Reads 4 original source Excel files and produces two standalone HTML files:
  1. org_drilldown.html        — full names
  2. org_drilldown_redacted.html — names replaced with redacted format

This is the backward-compatible generator. The default generator (generate_org_html.py)
reads from the single master Excel instead.
"""

import json
import re
import copy
import sys
from pathlib import Path
from collections import defaultdict

from org_html_shared import (
    normalize_name,
    slugify,
    is_contractor,
    title_seniority_score,
    redact_data,
    verify_redaction,
    generate_html,
    _HTML_TEMPLATE,
)

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ─── Configuration ───────────────────────────────────────────────────────────

DATA_DIR = Path(__file__).parent / "data"
LEGACY_DIR = DATA_DIR / "legacy"
ON24_FILE = LEGACY_DIR / "on24.xlsx"
ORG_FILE = LEGACY_DIR / "JayeshSahasi_QA-Dev Org List.xlsx"
SCRUMS_FILE = LEGACY_DIR / "JayeshSahasi_SCRUMS.xlsx"
TALENT_FILE = LEGACY_DIR / "JayeshSahasi_EngProduct_Jayesh_Talent_Snapshot_Leader_Input_2026.01.26.xlsx"
OUTPUT_FILE = Path(__file__).parent / "org_drilldown.html"
REDACTED_FILE = Path(__file__).parent / "org_drilldown_redacted.html"
MASTER_EXCEL_FILE = DATA_DIR / "orgchart_master_data.xlsx"

ORG_TABS = ["Product-Design", "Full QA Org", "Full Dev Org", "TPM", "ORG"]
QA_SUBTABS = ["Shefali", "Rumana", "Jenny", "QA Automation"]
SALESFORCE_ORG_NAME = "Salesforce"
TALENT_TABS = ["Dev", "QA", "Salesforce", "Product Management", "Program Management"]

JAYESH_NAME = "Jayesh Sahasi"
JAYESH_TITLE = "Executive VP, Products and CTO"
JAYESH_DRS = ["Jaimini", "Kamal", "Oleg", "Mahesh", "Steve Sims", "Jagjit"]

NICKNAME_MAP = {
    "steve": "stephen",
    "dan": "daniel",
    "jay": "jawynson",
    "mike": "michael",
    "ben": "benjamin",
    "raj": "raj",
}

# Full-name aliases: alternate full names → on24 canonical name (normalized)
NAME_ALIASES = {
    "jagjit singh": "jagjit bhullar",
    "mahesh khenny": "mahesh kheny",
    "jared chappins": "jared chappin",
    "ishwinder wallia": "ishwinder walia",
    "jose viscasillas": "jose viscasillas reyes",
    "michale lin": "michael lin",
    "rajesh kumar": "kumar rajesh",
    "bhagyashree more": "c-bhagyashree more",
}

# Kamal identity (on24.xlsx is definitive — reports to Jayesh)
KAMAL_FULL_NAME = "Kamal Ghosh"
KAMAL_TITLE = "Vice President Engineering"

# Manual title overrides for people whose names differ between files
MANUAL_TITLE_OVERRIDES = {
    "jagjit singh": "Director, Program Management",
    "jagjit bhullar": "Director, Program Management",
    "kamal ghosh": "Vice President Engineering",
}

# QA Org hierarchy fix: automation contractors report to Ashish, not directly to Oleg.
# Oleg's real direct reports are: Rumana, Shefali, Jenny, Ashish.
QA_OLEG_REAL_DRS = {"rumana", "shefali", "jenny", "ashish"}
QA_AUTOMATION_MANAGER = "ashish"  # Ashish Oza manages the automation team

# Default titles for people with no title found, by org tab
DEFAULT_TITLES_BY_ORG = {
    "Full Dev Org": "Senior Software Engineer",
    "Full QA Org": "Sr. QA Engineer",
    "Product-Design": "Senior UX Designer",
    "TPM": "Scrum Master",
}

# Per-person title overrides (normalized name -> title)
MANUAL_TITLE_OVERRIDES_EXTRA = {
    "sanel selimovic": "Senior UX Designer",
}

# DR-to-org mapping: which Jayesh DR heads which org view
DR_ORG_MAP = {
    "stephen sims": "Product-Design",
    "steve sims": "Product-Design",
    "oleg massakovskyy": "Full QA Org",
    "jaimini joshi": "Full Dev Org",
    "kamal ghosh": "Full Dev Org",
    "mahesh kheny": SALESFORCE_ORG_NAME,
    "jagjit bhullar": "TPM",
    "jagjit singh": "TPM",
}

# Changelog/notes rows that are NOT people — skip during parsing
CHANGELOG_SKIP_NAMES = {
    "issh", "update by", "notes", "created xl", "added teams", "date",
    "updated",
}

# Canonical scrum teams (from JayeshSahasi_SCRUMS.xlsx "Team Reorg" tab):
#  1. Analytics           2. Integration        3. Segmentation
#  4. Video               5. Cloud Engineering   6. Eng Tools
#  7. EHub/Target         8. VC                  9. GoLive
# 10. Engineering Support 11. Forums            12. Appgen
# 13. Console             14. Elite Studio      15. Elite Admin
# 16. Salesforce          17. Automation        18. Engineering AI
# 19. TPM
# Plus EER and Presenter (in org data but not in SCRUMS.xlsx)

TEAM_ALIASES = {
    # Console
    "p10console": "Console",
    "p10-console": "Console",
    "p10console forums": "Console",
    "p10-console forums": "Console",
    "p10console forums-": "Console",
    "console/es": "Console",
    "console": "Console",
    # Vids
    "vids": "Vids",
    "vids-vibbio": "Vids",
    "vids/vibbio": "Vids",
    "vids/forums": "Vids",
    "vids/es": "Vids",
    "video": "Vids",
    # Go Live
    "gl": "Go Live",
    "golive": "Go Live",
    "go live": "Go Live",
    "golife": "Go Live",
    # Elite Studio
    "es": "Elite Studio",
    "elite studio": "Elite Studio",
    "elite studio forums": "Elite Studio",
    "elite studio/forums": "Elite Studio",
    "elite studio/eng tools": "Elite Studio",
    "es forums": "Elite Studio",
    "es/forums": "Elite Studio",
    # Elite Admin
    "elite admin": "Elite Admin",
    "elite admin forums": "Elite Admin",
    "elite admin/elite admin forums": "Elite Admin",
    "elite admin/ai": "Elite Admin",
    "elite admin / wcc": "Elite Admin",
    # EHub/Target
    "ehub": "EHub/Target",
    "ehub/target": "EHub/Target",
    "ehub/target (orion)": "EHub/Target",
    # Integrations
    "integrations": "Integrations",
    "integration": "Integrations",
    # Engineering AI
    "ai": "Engineering AI",
    "engineering ai": "Engineering AI",
    "ai/analytics/integrations": "Engineering AI",
    "ai/integrations": "Engineering AI",
    # Cloud Engineering
    "cloud": "Cloud Engineering",
    "cloud engineering": "Cloud Engineering",
    "cloud engineering\n(ce)": "Cloud Engineering",
    # Eng Tools
    "engg tools": "Eng Tools",
    "eng tools": "Eng Tools",
    "tools": "Eng Tools",
    "eng tools /es forums": "Eng Tools",
    # Engineering Support
    "engg support": "Engineering Support",
    "eng support": "Engineering Support",
    "engineering support": "Engineering Support",
    "engineering support (cogs)": "Engineering Support",
    # Segmentation — NOT STAFFED, drop entirely
    "analytics -segmentation": None,
    "analytics-segmentation": None,
    "segmentation": None,
    # Appgen
    "appgen": "Appgen",
    # EER (not in SCRUMS.xlsx but referenced in Dev Org)
    "eer": "EER",
    "eer/gl": "EER",
    # Presenter (not in SCRUMS.xlsx but referenced in QA)
    "presenter": "Presenter",
    # Automation
    "automation": "Automation",
    # VC
    "vc": "VC",
    # Forums
    "forums": "Forums",
    "forums (kms)": "Forums",
    # Analytics
    "analytics": "Analytics",
    # Salesforce
    "salesforce": "Salesforce",
    # TPM
    "tpm": "TPM",
    # Non-scrum labels (skip these)
    "product mgmt": None,
    "ux": None,
    "ui/ux": None,
    "ui": None,
    "dir": None,
    "ft": None,
}

# Compound raw values that map to MULTIPLE teams (after % removal)
COMPOUND_TEAM_MAP = {
    "p10console forums- vids-vibbio-": ["Console", "Vids"],
    "p10console forums-  vids-vibbio-": ["Console", "Vids"],
}

# ─── Utilities ───────────────────────────────────────────────────────────────
# normalize_name, slugify, is_contractor, title_seniority_score imported from org_html_shared


def normalize_team_name(name):
    """Normalize a team name using aliases. Returns None to skip non-scrum labels."""
    key = name.strip().lower()
    key = re.sub(r'\s+', ' ', key)
    if key in TEAM_ALIASES:
        return TEAM_ALIASES[key]  # May be None for non-scrum labels
    # Keep as-is if not in aliases
    return name.strip()


def parse_scrum_teams(raw):
    """Parse team string into list of team names."""
    if not raw or not isinstance(raw, str):
        return []
    raw = raw.strip()
    if not raw or raw.lower() in ('nan', 'none', 'n/a', '-', '', 'dir'):
        return []
    # Remove percentage annotations
    raw = re.sub(r'\d+\s*%', '', raw)
    # Check compound team map first
    raw_key = raw.strip().lower()
    raw_key = re.sub(r'\s+', ' ', raw_key)
    if raw_key in COMPOUND_TEAM_MAP:
        return COMPOUND_TEAM_MAP[raw_key]
    # Split by common delimiters
    tokens = re.split(r'[,;/|&\n]+', raw)
    teams = []
    seen = set()
    for t in tokens:
        t = t.strip()
        t = re.sub(r'\s+', ' ', t)
        if not t or len(t) < 2:
            continue
        # Skip numeric-only tokens (changelog artifacts)
        if re.match(r'^\d+$', t):
            continue
        t = normalize_team_name(t)
        if t is None:
            continue  # Non-scrum label (Product Mgmt, UX, Dir, etc.)
        key = t.lower()
        if key not in seen:
            seen.add(key)
            teams.append(t)
    return teams


def is_changelog_row(reports_to, name):
    """Filter out changelog/audit rows at bottom of sheets."""
    if not name or not isinstance(name, str):
        return True
    name_lower = name.strip().lower()
    rt_lower = (reports_to or "").strip().lower() if isinstance(reports_to, str) else ""

    if name_lower in CHANGELOG_SKIP_NAMES or name_lower == "":
        return True
    # Date-like Reports To
    if rt_lower and re.match(r'^(date|dec|feb|jan|mar|apr|may|jun|jul|aug|sep|oct|nov)\b', rt_lower):
        return True
    # If name is just a number or very short
    if re.match(r'^\d+$', name.strip()):
        return True
    return False


# ─── Step 0a: Parse on24.xlsx — Definitive Hierarchy ────────────────────────

def convert_last_first(name_str):
    """Convert 'Last, First' to 'First Last'. Handles edge cases."""
    if not name_str or not isinstance(name_str, str):
        return ""
    name_str = name_str.strip()
    if ',' not in name_str:
        return name_str
    parts = name_str.split(',', 1)
    last = parts[0].strip()
    first = parts[1].strip()
    if not first or not last:
        return name_str
    return f"{first} {last}"


def parse_on24(filepath):
    """Parse on24.xlsx and return Jayesh's org subtree as a dict.

    Returns: {normalized_name -> {name, reports_to, title, department, location, num_drs}}
    Only includes people in Jayesh's subtree (BFS from Jayesh down).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["on24"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return {}

    header = [str(c).strip() if c else "" for c in rows[0]]
    header_lower = [h.lower() for h in header]

    # Find column indices
    col_map = {}
    for i, h in enumerate(header_lower):
        if h == "name":
            col_map["name"] = i
        elif h == "reports to":
            col_map["reports_to"] = i
        elif h == "job title":
            col_map["title"] = i
        elif h == "number of direct reports":
            col_map["num_drs"] = i
        elif h == "department":
            col_map["dept"] = i
        elif h == "location":
            col_map["loc"] = i

    if "name" not in col_map or "reports_to" not in col_map:
        print("  [WARN] on24.xlsx missing Name or Reports To column")
        return {}

    # Parse all rows — build name→record and reports_to index
    all_people = {}  # normalized_name -> record
    children_of = defaultdict(list)  # normalized_manager -> [normalized_name, ...]

    for row in rows[1:]:
        def cell(key):
            idx = col_map.get(key)
            if idx is not None and len(row) > idx and row[idx]:
                return str(row[idx]).strip()
            return ""

        raw_name = cell("name")
        raw_rt = cell("reports_to")
        if not raw_name:
            continue

        name = convert_last_first(raw_name)
        reports_to = convert_last_first(raw_rt)
        norm = normalize_name(name)
        norm_rt = normalize_name(reports_to)

        num_drs_raw = cell("num_drs")
        try:
            num_drs = int(num_drs_raw)
        except (ValueError, TypeError):
            num_drs = 0

        all_people[norm] = {
            "name": name,
            "reports_to": reports_to,
            "title": cell("title"),
            "department": cell("dept"),
            "location": cell("loc"),
            "num_drs": num_drs,
        }
        if norm_rt:
            children_of[norm_rt].append(norm)

    # BFS from Jayesh to collect his subtree
    jayesh_norm = normalize_name(JAYESH_NAME)
    if jayesh_norm not in all_people:
        print(f"  [WARN] '{JAYESH_NAME}' not found in on24.xlsx")
        return {}

    subtree = {}
    queue = [jayesh_norm]
    while queue:
        current = queue.pop(0)
        if current in subtree:
            continue
        if current not in all_people:
            continue
        subtree[current] = all_people[current]
        for child_norm in children_of.get(current, []):
            if child_norm not in subtree:
                queue.append(child_norm)

    print(f"  on24.xlsx: {len(all_people)} total people, {len(subtree)} in Jayesh's org")
    return subtree


# ─── Step 0a-2: Parse Contractor Locations ─────────────────────────────────

def parse_contractor_locations(filepath):
    """Parse SCRUMS.xlsx 'Contractor - Staff List' for country data.

    Returns: {normalized_name -> country_string}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = "Contractor - Staff List"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return {}

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    name_col = None
    country_col = None
    for i, h in enumerate(header):
        if h == "resource":
            name_col = i
        elif h == "country":
            country_col = i

    if name_col is None or country_col is None:
        return {}

    locations = {}
    for row in rows[1:]:
        raw_name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ""
        raw_country = str(row[country_col]).strip() if len(row) > country_col and row[country_col] else ""
        if not raw_name or raw_name.lower() in ("none", "nan", ""):
            continue
        if not raw_country or raw_country.lower() in ("none", "nan", ""):
            continue
        norm = normalize_name(raw_name)
        locations[norm] = raw_country
    return locations


# ─── Step 0b: Parse Teams Hierarchy ────────────────────────────────────────

def parse_teams_hierarchy(filepath):
    """Parse SCRUMS.xlsx 'Teams Hierachy' sheet for team→leader mappings.

    Returns: {canonical_team_name -> {dev_leads: [names], qa_leads: [names],
              dev_director: name, qa_director: name, vp: name}}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = "Teams Hierachy"  # Known typo in file
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Sheet '{sheet_name}' not found in SCRUMS.xlsx")
        wb.close()
        return {}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(rows) < 2:
        return {}

    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    # Find columns
    col_team = col_vp = col_dir = col_leads = col_qa_vp = col_qa_dir = col_qa_leads = None
    for i, h in enumerate(header):
        if h == "team":
            col_team = i
        elif h == "vp":
            col_vp = i
        elif h == "director":
            col_dir = i
        elif h in ("mgrs/leads", "mgrs / leads"):
            col_leads = i
        elif h == "qa vp":
            col_qa_vp = i
        elif h == "qa director":
            col_qa_dir = i
        elif h in ("qa leads", "qa lead"):
            col_qa_leads = i

    if col_team is None:
        print("  [WARN] Teams Hierarchy sheet missing 'Team' column")
        return {}

    def split_names(cell_val):
        """Split a cell value like 'Name1 / Name2' into list of names."""
        if not cell_val or not isinstance(cell_val, str):
            return []
        # Split on / and strip
        return [n.strip() for n in re.split(r'[/]', cell_val) if n.strip()]

    def cell_str(row, idx):
        if idx is not None and len(row) > idx and row[idx]:
            return str(row[idx]).strip()
        return ""

    teams_hier = {}
    for row in rows[1:]:
        team_raw = cell_str(row, col_team)
        if not team_raw:
            continue

        # Normalize team name
        team_name = normalize_team_name(team_raw)
        if team_name is None:
            continue

        teams_hier[team_name] = {
            "vp": cell_str(row, col_vp),
            "dev_director": cell_str(row, col_dir),
            "dev_leads": split_names(cell_str(row, col_leads)),
            "qa_director": cell_str(row, col_qa_dir),
            "qa_leads": split_names(cell_str(row, col_qa_leads)),
        }

    print(f"  Teams hierarchy: {len(teams_hier)} teams loaded")
    return teams_hier


# ─── Step 1: Parse Talent Snapshot ───────────────────────────────────────────

def parse_talent_snapshot(filepath):
    """Build title_map: normalize_name -> {title, talentBand, talentCategory, rationale}."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    title_map = {}

    for tab_name in TALENT_TABS:
        if tab_name not in wb.sheetnames:
            print(f"  [WARN] Talent tab '{tab_name}' not found, skipping")
            continue
        ws = wb[tab_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue

        # Find header row
        header = [str(c).strip() if c else "" for c in rows[0]]

        # Find column indices
        first_name_idx = None
        last_name_idx = None
        title_idx = None
        band_idx = None
        category_idx = None
        rationale_idx = None
        for i, h in enumerate(header):
            hl = h.lower()
            if hl == "first name":
                first_name_idx = i
            elif hl == "last name":
                last_name_idx = i
            elif hl == "title":
                title_idx = i
            elif "talent band" in hl:
                band_idx = i
            elif "talent category" in hl:
                category_idx = i
            elif "rationale" in hl:
                rationale_idx = i

        if first_name_idx is None or last_name_idx is None or title_idx is None:
            print(f"  [WARN] Talent tab '{tab_name}' missing required columns, skipping")
            continue

        max_needed = max(first_name_idx, last_name_idx, title_idx)
        for row in rows[1:]:
            if len(row) <= max_needed:
                continue
            first = str(row[first_name_idx]).strip() if row[first_name_idx] else ""
            last = str(row[last_name_idx]).strip() if row[last_name_idx] else ""
            title = str(row[title_idx]).strip() if row[title_idx] else ""

            if not first and not last:
                continue

            full = f"{first} {last}".strip()
            key = normalize_name(full)
            if key and title:
                # Clean up title
                title = re.sub(r'\s+', ' ', title).strip()

                band = ""
                if band_idx is not None and len(row) > band_idx and row[band_idx]:
                    band = str(row[band_idx]).strip()
                category = ""
                if category_idx is not None and len(row) > category_idx and row[category_idx]:
                    category = str(row[category_idx]).strip()
                rationale = ""
                if rationale_idx is not None and len(row) > rationale_idx and row[rationale_idx]:
                    rationale = re.sub(r'\s+', ' ', str(row[rationale_idx]).strip())

                title_map[key] = {
                    "title": title,
                    "talentBand": band,
                    "talentCategory": category,
                    "rationale": rationale,
                }

    wb.close()
    print(f"  Talent snapshot: {len(title_map)} entries loaded")
    return title_map


# ─── Step 2: Parse Org Roster ────────────────────────────────────────────────

def parse_org_tab(ws, tab_name):
    """Parse a single org tab for employment status and scrum teams only.

    Returns: dict of normalized_name -> {name, employment, scrumTeams, teamRaw,
             reportsToRaw, org_tab}
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return {}

    header = [str(c).strip() if c else "" for c in rows[0]]
    header_lower = [h.lower() for h in header]

    # Find column indices
    def find_col(candidates):
        for c in candidates:
            for i, h in enumerate(header_lower):
                if h == c.lower():
                    return i
        return None

    reports_to_idx = find_col(["Reports To", "reports to", "Managers", "managers"])
    name_idx = find_col(["Name", "name", "Employee Name", "employee name"])
    employment_idx = find_col(["Employment", "employment"])
    title_idx = find_col(["Title", "title"])

    # Team column: try Team, Teams, Teams.1 in priority
    team_idx = find_col(["Team", "team"])
    teams_idx = find_col(["Teams", "teams"])
    teams1_idx = find_col(["Teams.1", "teams.1"])

    if name_idx is None:
        print(f"  [WARN] Tab '{tab_name}' missing Name column")
        return {}

    people = {}
    last_reports_to = None

    for row in rows[1:]:
        if len(row) <= name_idx:
            continue

        raw_name = str(row[name_idx]).strip() if row[name_idx] else ""

        # Forward-fill Reports To (still needed for contractor manager resolution)
        raw_reports_to = ""
        if reports_to_idx is not None and len(row) > reports_to_idx and row[reports_to_idx]:
            raw_reports_to = str(row[reports_to_idx]).strip()
        if raw_reports_to and raw_reports_to.lower() not in ('none', 'nan', ''):
            last_reports_to = raw_reports_to
        else:
            raw_reports_to = last_reports_to or ""

        if not raw_name or raw_name.lower() in ('none', 'nan'):
            continue

        # Skip changelog rows
        if is_changelog_row(raw_reports_to, raw_name):
            continue

        # Employment
        employment = ""
        if employment_idx is not None and len(row) > employment_idx and row[employment_idx]:
            employment = str(row[employment_idx]).strip()
        if not employment:
            # Infer from c-prefix or Title column
            raw_title_val = ""
            if title_idx is not None and len(row) > title_idx and row[title_idx]:
                raw_title_val = str(row[title_idx]).strip()
            if raw_name.lower().startswith('c-'):
                employment = "Contractor"
            elif 'contract' in raw_title_val.lower():
                employment = "Contractor"
            else:
                employment = "Full Time"

        # Team(s)
        team_raw_parts = []
        for tidx in [team_idx, teams_idx, teams1_idx]:
            if tidx is not None and len(row) > tidx and row[tidx]:
                val = str(row[tidx]).strip()
                if val.lower() not in ('none', 'nan', ''):
                    team_raw_parts.append(val)
        team_raw = " / ".join(team_raw_parts) if team_raw_parts else ""
        scrum_teams = parse_scrum_teams(team_raw)

        norm = normalize_name(raw_name)
        people[norm] = {
            "name": raw_name,
            "employment": employment,
            "scrumTeams": scrum_teams,
            "teamRaw": team_raw,
            "reportsToRaw": raw_reports_to,
            "org_tab": tab_name,
        }

    return people


def fuzzy_title_match(name, title_map):
    """Try fuzzy matching for title lookup. Returns matched key or None."""
    norm = normalize_name(name)
    parts = norm.split()
    if len(parts) < 2:
        return None

    first = parts[0]
    last = parts[-1]

    # Try nickname
    first_alt = NICKNAME_MAP.get(first, first)

    for key in title_map:
        key_parts = key.split()
        if len(key_parts) < 2:
            continue
        k_first = key_parts[0]
        k_last = key_parts[-1]

        # Same last name + first name starts with or vice versa
        if k_last == last:
            if k_first.startswith(first) or first.startswith(k_first):
                return key
            if k_first.startswith(first_alt) or first_alt.startswith(k_first):
                return key

        # Reversed name
        if k_first == last and k_last == first:
            return key

    return None


def parse_org_roster(filepath):
    """Parse all org tabs for employment + teams. Returns merged dict of norm_name -> info."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_people = {}

    for tab_name in ORG_TABS:
        if tab_name not in wb.sheetnames:
            print(f"  [WARN] Org tab '{tab_name}' not found")
            continue
        ws = wb[tab_name]
        people = parse_org_tab(ws, tab_name)
        print(f"  {tab_name}: {len(people)} people")
        # Merge — later tabs don't overwrite earlier ones (first occurrence wins)
        for norm, info in people.items():
            if norm not in all_people:
                all_people[norm] = info
            else:
                # Merge missing fields from later tabs
                existing = all_people[norm]
                if info["scrumTeams"] and not existing["scrumTeams"]:
                    existing["scrumTeams"] = info["scrumTeams"]
                    existing["teamRaw"] = info["teamRaw"]
                if info["reportsToRaw"] and not existing["reportsToRaw"]:
                    existing["reportsToRaw"] = info["reportsToRaw"]

    wb.close()
    print(f"  Total from per-org tabs: {len(all_people)} unique people")
    return all_people


def parse_qa_subtabs(filepath):
    """Parse QA sub-hierarchy tabs (Shefali, Rumana, Jenny, QA Automation) for manager overrides and titles.

    Each tab has columns: Employee Name (A), Reports To (B), Title (C), Department (D).
    Returns: (overrides, titles) where:
      - overrides: dict of normalized_employee_name -> normalized_manager_name
      - titles: dict of normalized_employee_name -> title string
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    overrides = {}
    titles = {}

    for tab_name in QA_SUBTABS:
        if tab_name not in wb.sheetnames:
            print(f"  [WARN] QA subtab '{tab_name}' not found")
            continue
        ws = wb[tab_name]
        count = 0
        # Only read columns A-D (the primary data table), skip secondary mini-tables
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            raw_name = str(row[0]).strip() if row[0] else ""
            raw_mgr = str(row[1]).strip() if row[1] else ""
            raw_title = str(row[2]).strip() if row[2] else ""
            if not raw_name or raw_name.lower() in ('none', 'nan', ''):
                continue
            if not raw_mgr or raw_mgr.lower() in ('none', 'nan', ''):
                continue
            if is_changelog_row(raw_mgr, raw_name):
                continue
            norm_name = normalize_name(raw_name)
            norm_mgr = normalize_name(raw_mgr)
            overrides[norm_name] = norm_mgr
            # Capture title, stripping "Contractor" prefix/suffix
            if raw_title and raw_title.lower() not in ('none', 'nan', ''):
                clean_title = re.sub(r'\bContractor\b', '', raw_title, flags=re.IGNORECASE).strip()
                clean_title = clean_title.replace('\ufffd', '').strip()
                clean_title = re.sub(r'^[,\s\-\u2013\u2014]+|[,\s\-\u2013\u2014]+$', '', clean_title).strip()
                if clean_title:
                    titles[norm_name] = clean_title
            count += 1
        print(f"  QA subtab '{tab_name}': {count} reporting relationships")

    wb.close()
    print(f"  Total QA subtab overrides: {len(overrides)}, titles: {len(titles)}")
    return overrides, titles


# ─── Step 3: Build Tree Per Org ──────────────────────────────────────────────

def resolve_name_match(target, candidates_by_norm):
    """Resolve a name reference to a node. Returns node or None."""
    norm_target = normalize_name(target)

    # Exact match
    if norm_target in candidates_by_norm:
        return candidates_by_norm[norm_target]

    # Nickname substitution
    parts = norm_target.split()
    if parts:
        alt_first = NICKNAME_MAP.get(parts[0], parts[0])
        alt_name = " ".join([alt_first] + parts[1:])
        if alt_name in candidates_by_norm:
            return candidates_by_norm[alt_name]

    # Partial match: if target is single word, find by first-name contains
    if len(parts) == 1:
        matches = []
        for key, node in candidates_by_norm.items():
            if key.startswith(norm_target) or norm_target in key.split()[0]:
                matches.append(node)
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            # Pick highest seniority
            matches.sort(key=lambda n: title_seniority_score(n.get("title", "")), reverse=True)
            return matches[0]

    # Last name match + first name prefix
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        alt_first = NICKNAME_MAP.get(first, first)
        for key, node in candidates_by_norm.items():
            key_parts = key.split()
            if len(key_parts) >= 2:
                k_first, k_last = key_parts[0], key_parts[-1]
                if k_last == last and (k_first.startswith(first) or first.startswith(k_first) or
                                       k_first.startswith(alt_first) or alt_first.startswith(k_first)):
                    return node

    return None


def resolve_on24_name(target_name, on24_people):
    """Resolve a short/nickname reference to an on24 person. Returns norm key or None."""
    norm = normalize_name(target_name)
    if norm in on24_people:
        return norm

    # Full-name alias lookup
    alias = NAME_ALIASES.get(norm)
    if alias and alias in on24_people:
        return alias

    # Nickname substitution
    parts = norm.split()
    if parts:
        alt_first = NICKNAME_MAP.get(parts[0], parts[0])
        alt_name = " ".join([alt_first] + parts[1:])
        if alt_name in on24_people:
            return alt_name

    # Single word: partial first-name match
    if len(parts) == 1:
        matches = [k for k in on24_people if k.startswith(norm) or norm in k.split()[0]]
        if len(matches) == 1:
            return matches[0]

    # Last name + first name prefix
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        alt_first = NICKNAME_MAP.get(first, first)
        for key in on24_people:
            kp = key.split()
            if len(kp) >= 2:
                kf, kl = kp[0], kp[-1]
                if kl == last and (kf.startswith(first) or first.startswith(kf) or
                                   kf.startswith(alt_first) or alt_first.startswith(kf)):
                    return key
    return None


def build_from_on24(on24_people, org_tab_people, teams_hier, title_map, qa_subtab_overrides=None, qa_subtab_titles=None, contractor_locations=None):
    """Build all org datasets from on24 hierarchy + per-org enrichment.

    Returns: dict of org_name -> {top, nodes, children}
    """
    jayesh_norm = normalize_name(JAYESH_NAME)
    jayesh_id = slugify(JAYESH_NAME)

    # ── Phase 1: Create nodes from on24 data ──
    all_nodes = {}  # id -> node dict
    norm_to_id = {}  # normalized name -> node id
    children = defaultdict(list)

    for norm, person in on24_people.items():
        node_id = slugify(person["name"])
        # Avoid slug collisions
        if node_id in all_nodes:
            node_id = f"{node_id}-2"

        # Look up employment & scrum teams from per-org tabs
        tab_info = _match_org_tab(norm, person["name"], org_tab_people)

        node = {
            "id": node_id,
            "name": person["name"],
            "title": person["title"],
            "employment": tab_info.get("employment", "") if tab_info else "Full Time",
            "teamRaw": tab_info.get("teamRaw", "") if tab_info else "",
            "scrumTeams": tab_info.get("scrumTeams", []) if tab_info else [],
            "managerId": None,
            "placeholder": False,
            "org": "",  # assigned later
            "location": person.get("location", ""),
        }
        # Default employment for on24 FTEs
        if not node["employment"]:
            node["employment"] = "Full Time"

        all_nodes[node_id] = node
        norm_to_id[norm] = node_id

    # Override Jayesh title
    if jayesh_norm in norm_to_id:
        all_nodes[norm_to_id[jayesh_norm]]["title"] = JAYESH_TITLE
        jayesh_id = norm_to_id[jayesh_norm]

    # ── Phase 2: Resolve manager IDs from on24 hierarchy ──
    for norm, person in on24_people.items():
        node_id = norm_to_id[norm]
        if node_id == jayesh_id:
            all_nodes[node_id]["managerId"] = None
            continue
        rt_norm = normalize_name(person["reports_to"])
        if rt_norm in norm_to_id:
            mgr_id = norm_to_id[rt_norm]
            all_nodes[node_id]["managerId"] = mgr_id
            children[mgr_id].append(node_id)
        else:
            # Orphan in on24 — assign to Jayesh
            all_nodes[node_id]["managerId"] = jayesh_id
            children[jayesh_id].append(node_id)

    # ── Phase 2b: Enrich on24 nodes with talent snapshot data ──
    for norm, node_id in norm_to_id.items():
        talent = title_map.get(norm)
        if not talent:
            matched_key = fuzzy_title_match(norm, title_map)
            if matched_key:
                talent = title_map[matched_key]
        if talent:
            all_nodes[node_id]["talentBand"] = talent.get("talentBand", "")
            all_nodes[node_id]["talentCategory"] = talent.get("talentCategory", "")
            all_nodes[node_id]["rationale"] = talent.get("rationale", "")

    # ── Phase 3: Add people from per-org tabs not already in on24 ──
    slug_counter = defaultdict(int)
    on24_norms = set(norm_to_id.keys())

    for norm, info in org_tab_people.items():
        if norm in on24_norms:
            continue  # Already in on24 tree
        # Alias check: known alternate spellings
        alias = NAME_ALIASES.get(norm)
        if alias and (alias in norm_to_id or alias in on24_norms):
            continue
        # Fuzzy check: short org-tab names may match longer on24 names
        resolved = resolve_on24_name(info["name"], on24_people)
        if resolved and resolved in norm_to_id:
            continue  # Same person under a different name in on24

        name = info["name"]
        base_slug = slugify(name)
        slug_counter[base_slug] += 1
        node_id = base_slug if slug_counter[base_slug] == 1 else f"{base_slug}-{slug_counter[base_slug]}"
        # Avoid collision with on24 nodes
        while node_id in all_nodes:
            slug_counter[base_slug] += 1
            node_id = f"{base_slug}-{slug_counter[base_slug]}"

        # Resolve title: manual overrides > talent snapshot > fuzzy > org default
        title = ""
        talent_key = None  # track which title_map key matched
        if norm in MANUAL_TITLE_OVERRIDES:
            title = MANUAL_TITLE_OVERRIDES[norm]
            talent_key = norm if norm in title_map else None
        elif norm in MANUAL_TITLE_OVERRIDES_EXTRA:
            title = MANUAL_TITLE_OVERRIDES_EXTRA[norm]
            talent_key = norm if norm in title_map else None
        elif norm in title_map:
            title = title_map[norm]["title"]
            talent_key = norm
        else:
            matched_key = fuzzy_title_match(name, title_map)
            if matched_key:
                title = title_map[matched_key]["title"]
                talent_key = matched_key
            else:
                title = DEFAULT_TITLES_BY_ORG.get(info.get("org_tab", ""), "")

        # Attach talent snapshot fields
        talent_info = title_map.get(talent_key) if talent_key else None
        node = {
            "id": node_id,
            "name": name,
            "title": title,
            "employment": info.get("employment", ""),
            "teamRaw": info.get("teamRaw", ""),
            "scrumTeams": info.get("scrumTeams", []),
            "managerId": None,
            "placeholder": False,
            "org": "",
            "location": "",
            "talentBand": talent_info["talentBand"] if talent_info else "",
            "talentCategory": talent_info["talentCategory"] if talent_info else "",
            "rationale": talent_info["rationale"] if talent_info else "",
        }

        # Resolve manager: try team lead from Teams Hierarchy, then per-org Reports To
        mgr_id = _resolve_contractor_manager(
            info, norm_to_id, on24_people, org_tab_people, teams_hier, all_nodes
        )
        node["managerId"] = mgr_id or jayesh_id
        all_nodes[node_id] = node
        norm_to_id[norm] = node_id
        if node["managerId"]:
            children[node["managerId"]].append(node_id)

    # ── Phase 4: QA org fix — automation contractors under Ashish ──
    oleg_norm = resolve_on24_name("Oleg", on24_people)
    ashish_norm = _match_org_tab_norm("ashish", org_tab_people)
    if oleg_norm and ashish_norm:
        oleg_id = norm_to_id.get(oleg_norm)
        ashish_id = norm_to_id.get(ashish_norm)
        if oleg_id and not ashish_id:
            # Ashish is a contractor not in on24 — create node if needed
            ashish_info = org_tab_people.get(ashish_norm)
            if ashish_info and ashish_norm in norm_to_id:
                ashish_id = norm_to_id[ashish_norm]
        if oleg_id and ashish_id:
            # Ensure Ashish reports to Oleg
            all_nodes[ashish_id]["managerId"] = oleg_id
            if ashish_id not in children.get(oleg_id, []):
                children[oleg_id].append(ashish_id)
            # Reroute non-real DRs of Oleg to Ashish
            oleg_children = list(children.get(oleg_id, []))
            for cid in oleg_children:
                if cid == ashish_id:
                    continue
                child_norm = normalize_name(all_nodes[cid]["name"])
                first_name = child_norm.split()[0] if child_norm else ""
                if first_name not in QA_OLEG_REAL_DRS:
                    all_nodes[cid]["managerId"] = ashish_id
                    children[oleg_id].remove(cid)
                    children[ashish_id].append(cid)

    # ── Phase 4b: QA subtab overrides — fix sub-hierarchies from Shefali/Rumana/Jenny tabs ──
    if qa_subtab_overrides:
        reparent_count = 0
        for person_norm, desired_mgr_norm in qa_subtab_overrides.items():
            # Resolve person to node ID
            person_id = norm_to_id.get(person_norm)
            if not person_id:
                # Try alias lookup
                alias = NAME_ALIASES.get(person_norm)
                if alias:
                    person_id = norm_to_id.get(alias)
            if not person_id:
                # Try fuzzy resolve (e.g. "c-" prefix variations)
                resolved = resolve_on24_name(person_norm, on24_people)
                if resolved:
                    person_id = norm_to_id.get(resolved)
                if not person_id:
                    # Try org tab match
                    resolved = _match_org_tab_norm(person_norm, org_tab_people)
                    if resolved:
                        person_id = norm_to_id.get(resolved)
            if not person_id:
                continue

            # Resolve desired manager to node ID
            mgr_id = norm_to_id.get(desired_mgr_norm)
            if not mgr_id:
                alias = NAME_ALIASES.get(desired_mgr_norm)
                if alias:
                    mgr_id = norm_to_id.get(alias)
            if not mgr_id:
                resolved = resolve_on24_name(desired_mgr_norm, on24_people)
                if resolved:
                    mgr_id = norm_to_id.get(resolved)
                if not mgr_id:
                    resolved = _match_org_tab_norm(desired_mgr_norm, org_tab_people)
                    if resolved:
                        mgr_id = norm_to_id.get(resolved)
            if not mgr_id:
                continue

            # Skip if already correctly parented
            current_mgr = all_nodes[person_id].get("managerId")
            if current_mgr == mgr_id:
                continue

            # Re-parent: remove from old parent's children, add to new
            if current_mgr and person_id in children.get(current_mgr, []):
                children[current_mgr].remove(person_id)
            all_nodes[person_id]["managerId"] = mgr_id
            if person_id not in children.get(mgr_id, []):
                children[mgr_id].append(person_id)
            reparent_count += 1

        print(f"  Phase 4b: {reparent_count} QA subtab re-parents applied")

    # ── Phase 4c: Apply titles from QA subtabs ──
    if qa_subtab_titles:
        title_count = 0
        for person_norm, subtab_title in qa_subtab_titles.items():
            person_id = norm_to_id.get(person_norm)
            if not person_id:
                alias = NAME_ALIASES.get(person_norm)
                if alias:
                    person_id = norm_to_id.get(alias)
            if not person_id:
                resolved = resolve_on24_name(person_norm, on24_people)
                if resolved:
                    person_id = norm_to_id.get(resolved)
                if not person_id:
                    resolved = _match_org_tab_norm(person_norm, org_tab_people)
                    if resolved:
                        person_id = norm_to_id.get(resolved)
            if not person_id:
                continue
            # Skip if person has a manual override or talent snapshot title
            if person_norm in MANUAL_TITLE_OVERRIDES or person_norm in MANUAL_TITLE_OVERRIDES_EXTRA:
                continue
            node = all_nodes[person_id]
            current_title = node.get("title", "")
            # Apply subtab title if person has no title, a default title, or a generic org default
            if not current_title or current_title in DEFAULT_TITLES_BY_ORG.values():
                node["title"] = subtab_title
                title_count += 1
        print(f"  Phase 4c: {title_count} QA subtab titles applied")

    # ── Phase 4d: Fill contractor locations from SCRUMS contractor list ──
    if contractor_locations:
        loc_count = 0
        for norm, country in contractor_locations.items():
            nid = norm_to_id.get(norm)
            if not nid:
                alias = NAME_ALIASES.get(norm)
                if alias:
                    nid = norm_to_id.get(alias)
            if nid and not all_nodes[nid].get("location"):
                all_nodes[nid]["location"] = country
                loc_count += 1
        print(f"  Phase 4d: {loc_count} contractor locations applied")

    # ── Phase 5: Assign orgs based on DR subtree ──
    # Build DR -> org mapping
    dr_org = {}  # node_id -> org_name
    for dr_hint in JAYESH_DRS:
        dr_norm = resolve_on24_name(dr_hint, on24_people)
        if not dr_norm:
            dr_norm = _match_org_tab_norm(normalize_name(dr_hint), org_tab_people)
        if dr_norm and dr_norm in norm_to_id:
            dr_id = norm_to_id[dr_norm]
            # Determine org from DR_ORG_MAP
            for name_key, org_name in DR_ORG_MAP.items():
                if name_key == dr_norm or dr_norm.startswith(name_key.split()[0]):
                    dr_org[dr_id] = org_name
                    break

    # BFS from each DR to assign org to entire subtree
    node_orgs = {}  # node_id -> org_name
    node_orgs[jayesh_id] = "Home"

    for dr_id, org_name in dr_org.items():
        queue = [dr_id]
        while queue:
            nid = queue.pop(0)
            if nid in node_orgs:
                continue
            node_orgs[nid] = org_name
            for child_id in children.get(nid, []):
                queue.append(child_id)

    # Assign org to nodes
    for nid, node in all_nodes.items():
        node["org"] = node_orgs.get(nid, "Full Dev Org")

    # ── Phase 6: Split into per-org datasets ──
    org_datasets = {}
    for org_name in list(dict.fromkeys(
        ["Product-Design", "Full QA Org", "Full Dev Org", SALESFORCE_ORG_NAME, "TPM"]
    )):
        org_node_ids = {nid for nid, o in node_orgs.items() if o == org_name}
        if not org_node_ids:
            continue

        # Always include Jayesh at top
        org_nodes = {jayesh_id: all_nodes[jayesh_id].copy()}
        org_nodes[jayesh_id]["org"] = org_name
        org_children = defaultdict(list)

        # Jayesh's DRs for this org
        for dr_id, dr_org_name in dr_org.items():
            if dr_org_name == org_name:
                org_children[jayesh_id].append(dr_id)

        # Add all nodes in this org
        for nid in org_node_ids:
            org_nodes[nid] = all_nodes[nid]
            # Add children that are also in this org
            for cid in children.get(nid, []):
                if cid in org_node_ids:
                    org_children[nid].append(cid)

        # Sort children alphabetically
        for pid in org_children:
            org_children[pid].sort(
                key=lambda cid: org_nodes.get(cid, all_nodes.get(cid, {})).get("name", "").lower()
            )

        org_datasets[org_name] = {
            "top": jayesh_id,
            "nodes": org_nodes,
            "children": dict(org_children),
        }

        # Report
        dr_names = [org_nodes.get(d, {}).get("name", "?") for d in org_children.get(jayesh_id, [])]
        print(f"  {org_name}: {len(org_nodes)} nodes, DRs: {dr_names}")

    return org_datasets


def _match_org_tab(norm, name, org_tab_people):
    """Find a person in org_tab_people by normalized name, with fuzzy fallback."""
    if norm in org_tab_people:
        return org_tab_people[norm]

    # Try nickname substitution
    parts = norm.split()
    if parts:
        alt_first = NICKNAME_MAP.get(parts[0], parts[0])
        alt_name = " ".join([alt_first] + parts[1:])
        if alt_name in org_tab_people:
            return org_tab_people[alt_name]

    # Partial: last name + first name prefix
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        alt_first = NICKNAME_MAP.get(first, first)
        for key, info in org_tab_people.items():
            kp = key.split()
            if len(kp) >= 2:
                kf, kl = kp[0], kp[-1]
                if kl == last and (kf.startswith(first) or first.startswith(kf) or
                                   kf.startswith(alt_first) or alt_first.startswith(kf)):
                    return info

    return None


def _match_org_tab_norm(target_norm, org_tab_people):
    """Find normalized key in org_tab_people, with fuzzy matching. Returns norm key."""
    if target_norm in org_tab_people:
        return target_norm
    parts = target_norm.split()
    if parts:
        alt_first = NICKNAME_MAP.get(parts[0], parts[0])
        alt_name = " ".join([alt_first] + parts[1:])
        if alt_name in org_tab_people:
            return alt_name
    # Partial first-name match
    if len(parts) == 1:
        matches = [k for k in org_tab_people if k.startswith(target_norm)]
        if len(matches) == 1:
            return matches[0]
    return None


def _resolve_contractor_manager(info, norm_to_id, on24_people, org_tab_people, teams_hier, all_nodes):
    """Resolve which on24 person a contractor should report to.

    Priority: team lead from Teams Hierarchy > per-org Reports To.
    """
    # Try Teams Hierarchy: find lead for contractor's scrum team
    for team in info.get("scrumTeams", []):
        th = teams_hier.get(team)
        if not th:
            continue
        # Determine discipline from org_tab
        org_tab = info.get("org_tab", "")
        if "QA" in org_tab:
            lead_names = th.get("qa_leads", [])
            if not lead_names:
                lead_names = [th.get("qa_director", "")]
        else:
            lead_names = th.get("dev_leads", [])
            if not lead_names:
                lead_names = [th.get("dev_director", "")]

        for lead_name in lead_names:
            if not lead_name:
                continue
            lead_norm = resolve_on24_name(lead_name, on24_people)
            if lead_norm and lead_norm in norm_to_id:
                return norm_to_id[lead_norm]

    # Fallback: use Reports To from per-org tab
    rt = info.get("reportsToRaw", "")
    if rt:
        rt_norm = normalize_name(rt)
        # Try direct match in norm_to_id
        if rt_norm in norm_to_id:
            return norm_to_id[rt_norm]
        # Try on24 resolve
        resolved = resolve_on24_name(rt, on24_people)
        if resolved and resolved in norm_to_id:
            return norm_to_id[resolved]
        # Try org_tab resolve (for contractor→contractor chains)
        tab_norm = _match_org_tab_norm(rt_norm, org_tab_people)
        if tab_norm and tab_norm in norm_to_id:
            return norm_to_id[tab_norm]

    return None


# ─── Step 4: Build Scrum Index ───────────────────────────────────────────────

def build_scrum_index(org_datasets, teams_hier=None):
    """Build global scrum index: team_name -> list of members.

    Uses teams_hier (from SCRUMS.xlsx) to explicitly identify leads.
    """
    teams_hier = teams_hier or {}

    # Build lead name sets per team for quick lookup
    team_lead_names = {}  # team_name_lower -> set of normalized lead names
    for team_name, th in teams_hier.items():
        lead_norms = set()
        for lead in th.get("dev_leads", []) + th.get("qa_leads", []):
            if lead:
                lead_norms.add(normalize_name(lead))
        team_lead_names[team_name.lower()] = lead_norms

    # First pass: collect all teams with case-insensitive dedup
    canonical_names = {}
    raw_index = defaultdict(list)

    for tab_name, dataset in org_datasets.items():
        for node_id, node in dataset["nodes"].items():
            if node.get("placeholder"):
                continue
            for team in node.get("scrumTeams", []):
                if team.lower() in ("dir", "management", ""):
                    continue
                key = team.lower().strip()
                if key not in canonical_names:
                    canonical_names[key] = team

                # Check if this person is a lead for this team
                person_norm = normalize_name(node["name"])
                person_first = person_norm.split()[0] if person_norm else ""
                lead_set = team_lead_names.get(key, set())
                is_lead = (person_norm in lead_set or
                           any(person_first and ln.startswith(person_first) for ln in lead_set))

                raw_index[key].append({
                    "org": tab_name,
                    "id": node_id,
                    "name": node["name"],
                    "title": node.get("title", ""),
                    "employment": node.get("employment", ""),
                    "isLead": is_lead,
                    "talentBand": node.get("talentBand", ""),
                    "talentCategory": node.get("talentCategory", ""),
                    "rationale": node.get("rationale", ""),
                })

    # Build canonical scrum_index
    scrum_index = {}
    for key, members in raw_index.items():
        display = canonical_names[key]
        scrum_index[display] = members

    # Determine discipline and sort (leads first)
    scrum_teams = {}
    for team_name, members in scrum_index.items():
        grouped = {"Dev": [], "QA": [], "Product": [], "TPM": [], "Salesforce": [], "Other": []}
        for m in members:
            org = m["org"]
            if org == SALESFORCE_ORG_NAME:
                grouped["Salesforce"].append(m)
            elif "Dev" in org:
                grouped["Dev"].append(m)
            elif "QA" in org:
                grouped["QA"].append(m)
            elif "Product" in org or "Design" in org:
                grouped["Product"].append(m)
            elif "TPM" in org:
                grouped["TPM"].append(m)
            else:
                grouped["Other"].append(m)

        # Sort: leads first, then by seniority
        for discipline in grouped:
            group = grouped[discipline]
            if group:
                group.sort(key=lambda m: (
                    0 if m.get("isLead") else 1,
                    -title_seniority_score(m["title"]),
                    m["name"].lower()
                ))

        scrum_teams[team_name] = grouped

    return scrum_teams


# ─── Step 5: HTML Generation ─────────────────────────────────────────────────

def make_serializable(org_datasets, scrum_teams):
    """Convert data to JSON-serializable format."""
    orgs = {}
    # Build Home view: Jayesh + his 6 DRs across all orgs (deduplicated)
    home_drs = []
    seen_dr_names = set()

    for tab_name, dataset in org_datasets.items():
        nodes_ser = {}
        top_id = dataset["top"]
        top_children = dataset["children"].get(top_id, [])

        for nid, node in dataset["nodes"].items():
            ser = {
                "id": node["id"],
                "name": node["name"],
                "title": node.get("title", ""),
                "employment": node.get("employment", ""),
                "scrumTeams": node.get("scrumTeams", []),
                "placeholder": node.get("placeholder", False),
                "org": node.get("org", tab_name),
                "talentBand": node.get("talentBand", ""),
                "talentCategory": node.get("talentCategory", ""),
                "rationale": node.get("rationale", ""),
            }
            if node.get("dottedLine"):
                ser["dottedLine"] = node["dottedLine"]
            nodes_ser[nid] = ser

        orgs[tab_name] = {
            "top": top_id,
            "nodes": nodes_ser,
            "children": dataset["children"],
        }

        # Collect Jayesh's DRs for the Home view (dedup across orgs)
        for cid in top_children:
            c = nodes_ser.get(cid)
            if c:
                dr_norm = normalize_name(c["name"])
                if dr_norm not in seen_dr_names:
                    seen_dr_names.add(dr_norm)
                    home_drs.append({
                        "name": c["name"],
                        "title": c.get("title", ""),
                        "employment": c.get("employment", ""),
                        "org": tab_name,
                        "nodeId": cid,
                        "talentBand": c.get("talentBand", ""),
                        "talentCategory": c.get("talentCategory", ""),
                        "rationale": c.get("rationale", ""),
                    })

    return {
        "orgs": orgs,
        "scrum": scrum_teams,
        "missing": {},
        "homeDrs": home_drs,
    }


## redact_data, verify_redaction, generate_html, _HTML_TEMPLATE
## are imported from org_html_shared


# ─── Step 8: Export Master Excel ────────────────────────────────────────────

def export_master_excel(org_datasets, scrum_teams, teams_hier):
    """Export all org chart data to a single master Excel workbook.

    Creates orgchart_master_data.xlsx with:
      - People sheet (deduped roster)
      - Scrum Teams sheet (person-team assignments)
      - Teams Hierarchy sheet (team leadership)
      - _Validation hidden sheet (dropdown lists)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── Collect all unique people across orgs ──
    seen_names = set()
    people = []
    for org_name, dataset in org_datasets.items():
        for nid, node in dataset["nodes"].items():
            if node.get("placeholder"):
                continue
            name_norm = normalize_name(node["name"])
            if name_norm in seen_names:
                continue
            seen_names.add(name_norm)

            # Resolve manager name from managerId
            mgr_name = ""
            mgr_id = node.get("managerId")
            if mgr_id:
                # Search across all org datasets for the manager node
                for ds in org_datasets.values():
                    mgr_node = ds["nodes"].get(mgr_id)
                    if mgr_node:
                        mgr_name = mgr_node["name"]
                        break

            people.append({
                "name": node["name"],
                "title": node.get("title", ""),
                "employment": node.get("employment", ""),
                "org": node.get("org", org_name),
                "reportsTo": mgr_name,
                "scrumTeams": "; ".join(node.get("scrumTeams", [])),
                "location": node.get("location", ""),
                "talentBand": node.get("talentBand", ""),
                "talentCategory": node.get("talentCategory", ""),
                "rationale": node.get("rationale", ""),
            })

    # Sort: by Org, then Reports To, then Name
    org_order = ["Full Dev Org", "Full QA Org", "Product-Design", "Salesforce", "TPM"]
    people.sort(key=lambda p: (
        org_order.index(p["org"]) if p["org"] in org_order else 99,
        p["reportsTo"].lower(),
        p["name"].lower(),
    ))

    # ── Collect known titles for validation ──
    known_titles = sorted({p["title"] for p in people if p["title"]})

    # ── Collect known talent band / category values ──
    talent_bands = sorted({p["talentBand"] for p in people if p["talentBand"]})
    talent_categories = sorted({p["talentCategory"] for p in people if p["talentCategory"]})

    # ── Canonical team list ──
    canonical_teams = sorted({
        t for t in TEAM_ALIASES.values() if t is not None
    })

    # ── Styling constants ──
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    def style_header(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.freeze_panes = "A2"

    def auto_width(ws, min_width=10, max_width=50):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = min_width
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
            ws.column_dimensions[col_letter].width = max_len + 2

    def apply_alt_rows(ws, start_row, end_row):
        for row_idx in range(start_row, end_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            for cell in ws[row_idx]:
                cell.border = thin_border

    # ── Hidden Sheet: _Validation ──
    ws_val = wb.active
    ws_val.title = "_Validation"

    known_locations = sorted({p["location"] for p in people if p["location"]})

    val_lists = {
        "valid_employment": ["Full Time", "Contractor"],
        "valid_orgs": org_order,
        "valid_teams": canonical_teams,
        "valid_locations": known_locations if known_locations else ["—"],
        "valid_bands": talent_bands if talent_bands else ["—"],
        "valid_categories": talent_categories if talent_categories else ["—"],
        "valid_disciplines": ["Dev", "QA", "Product", "TPM", "Salesforce", "Other"],
        "valid_yesno": ["Yes", "No"],
        "valid_titles": known_titles if known_titles else ["—"],
    }

    col = 1
    range_refs = {}  # name -> "$A$2:$A$N" style ref
    for list_name, values in val_lists.items():
        ws_val.cell(row=1, column=col, value=list_name).font = Font(bold=True)
        for row_idx, val in enumerate(values, 2):
            ws_val.cell(row=row_idx, column=col, value=val)
        col_letter = get_column_letter(col)
        range_refs[list_name] = f"'_Validation'!${col_letter}$2:${col_letter}${len(values) + 1}"
        col += 1

    ws_val.sheet_state = "hidden"

    # ── Sheet 1: People ──
    ws_people = wb.create_sheet("People", 0)
    people_headers = [
        "Name", "Title", "Employment", "Org", "Reports To",
        "Scrum Teams", "Location", "Talent Band", "Talent Category", "Rationale",
    ]
    style_header(ws_people, people_headers)

    for row_idx, p in enumerate(people, 2):
        ws_people.cell(row=row_idx, column=1, value=p["name"])
        ws_people.cell(row=row_idx, column=2, value=p["title"])
        ws_people.cell(row=row_idx, column=3, value=p["employment"])
        ws_people.cell(row=row_idx, column=4, value=p["org"])
        ws_people.cell(row=row_idx, column=5, value=p["reportsTo"])
        ws_people.cell(row=row_idx, column=6, value=p["scrumTeams"])
        ws_people.cell(row=row_idx, column=7, value=p["location"])
        ws_people.cell(row=row_idx, column=8, value=p["talentBand"])
        ws_people.cell(row=row_idx, column=9, value=p["talentCategory"])
        ws_people.cell(row=row_idx, column=10, value=p["rationale"])

    people_end = len(people) + 1
    apply_alt_rows(ws_people, 2, people_end)
    auto_width(ws_people)
    # Rationale column wider (col J)
    ws_people.column_dimensions["J"].width = 60
    # Auto-filter on all columns
    ws_people.auto_filter.ref = f"A1:{get_column_letter(len(people_headers))}{people_end}"

    # Data validation on People sheet
    dv_title = DataValidation(type="list", formula1=range_refs["valid_titles"], allow_blank=True)
    dv_title.prompt = "Select or type a title"
    dv_title.showErrorMessage = False  # Allow free text
    ws_people.add_data_validation(dv_title)
    dv_title.add(f"B2:B{people_end}")

    dv_emp = DataValidation(type="list", formula1=range_refs["valid_employment"], allow_blank=True)
    ws_people.add_data_validation(dv_emp)
    dv_emp.add(f"C2:C{people_end}")

    dv_org = DataValidation(type="list", formula1=range_refs["valid_orgs"], allow_blank=True)
    ws_people.add_data_validation(dv_org)
    dv_org.add(f"D2:D{people_end}")

    dv_loc = DataValidation(type="list", formula1=range_refs["valid_locations"], allow_blank=True)
    dv_loc.showErrorMessage = False  # Allow free text
    ws_people.add_data_validation(dv_loc)
    dv_loc.add(f"G2:G{people_end}")

    dv_band = DataValidation(type="list", formula1=range_refs["valid_bands"], allow_blank=True)
    ws_people.add_data_validation(dv_band)
    dv_band.add(f"H2:H{people_end}")

    dv_cat = DataValidation(type="list", formula1=range_refs["valid_categories"], allow_blank=True)
    ws_people.add_data_validation(dv_cat)
    dv_cat.add(f"I2:I{people_end}")

    # ── Sheet 2: Scrum Teams ──
    ws_scrum = wb.create_sheet("Scrum Teams")
    scrum_headers = ["Team Name", "Member Name", "Discipline", "Is Lead"]
    style_header(ws_scrum, scrum_headers)

    scrum_row = 2
    for team_name in sorted(scrum_teams.keys()):
        groups = scrum_teams[team_name]
        for discipline, members in groups.items():
            for m in members:
                ws_scrum.cell(row=scrum_row, column=1, value=team_name)
                ws_scrum.cell(row=scrum_row, column=2, value=m["name"])
                ws_scrum.cell(row=scrum_row, column=3, value=discipline)
                ws_scrum.cell(row=scrum_row, column=4, value="Yes" if m.get("isLead") else "No")
                scrum_row += 1

    scrum_end = scrum_row - 1
    if scrum_end >= 2:
        apply_alt_rows(ws_scrum, 2, scrum_end)
    auto_width(ws_scrum)
    # Auto-filter on all columns
    ws_scrum.auto_filter.ref = f"A1:{get_column_letter(len(scrum_headers))}{scrum_end}"

    dv_team = DataValidation(type="list", formula1=range_refs["valid_teams"], allow_blank=True)
    dv_team.showErrorMessage = False  # Allow non-canonical teams
    ws_scrum.add_data_validation(dv_team)
    dv_team.add(f"A2:A{scrum_end}")

    dv_disc = DataValidation(type="list", formula1=range_refs["valid_disciplines"], allow_blank=True)
    ws_scrum.add_data_validation(dv_disc)
    dv_disc.add(f"C2:C{scrum_end}")

    dv_lead = DataValidation(type="list", formula1=range_refs["valid_yesno"], allow_blank=True)
    ws_scrum.add_data_validation(dv_lead)
    dv_lead.add(f"D2:D{scrum_end}")

    # ── Sheet 3: Teams Hierarchy ──
    ws_th = wb.create_sheet("Teams Hierarchy")
    th_headers = ["Team Name", "Dev Lead(s)", "QA Lead(s)", "Dev Director", "QA Director"]
    style_header(ws_th, th_headers)

    th_row = 2
    for team_name in sorted(teams_hier.keys()):
        th = teams_hier[team_name]
        ws_th.cell(row=th_row, column=1, value=team_name)
        ws_th.cell(row=th_row, column=2, value=", ".join(th.get("dev_leads", [])))
        ws_th.cell(row=th_row, column=3, value=", ".join(th.get("qa_leads", [])))
        ws_th.cell(row=th_row, column=4, value=th.get("dev_director", ""))
        ws_th.cell(row=th_row, column=5, value=th.get("qa_director", ""))
        th_row += 1

    th_end = th_row - 1
    if th_end >= 2:
        apply_alt_rows(ws_th, 2, th_end)
    auto_width(ws_th)
    # Auto-filter on all columns
    ws_th.auto_filter.ref = f"A1:{get_column_letter(len(th_headers))}{th_end}"

    dv_team_th = DataValidation(type="list", formula1=range_refs["valid_teams"], allow_blank=True)
    dv_team_th.showErrorMessage = False
    ws_th.add_data_validation(dv_team_th)
    dv_team_th.add(f"A2:A{th_end}")

    # ── Save ──
    wb.save(MASTER_EXCEL_FILE)
    print(f"  Written: {MASTER_EXCEL_FILE}")
    print(f"    People: {len(people)} rows")
    print(f"    Scrum Teams: {scrum_end - 1} rows")
    print(f"    Teams Hierarchy: {th_end - 1} rows")
    print(f"    Titles: {len(known_titles)}, Locations: {len(known_locations)}, Bands: {len(talent_bands)}, Categories: {len(talent_categories)}")


def main():
    print("=" * 60)
    print("Org Chart HTML Generator (on24.xlsx hierarchy)")
    print("=" * 60)

    # Step 1: Parse on24.xlsx — definitive FTE hierarchy
    print("\n[1] Parsing on24.xlsx (definitive hierarchy)...")
    on24_people = parse_on24(ON24_FILE)

    # Step 2: Parse talent snapshot — supplementary titles
    print("\n[2] Parsing talent snapshot...")
    title_map = parse_talent_snapshot(TALENT_FILE)

    # Step 3: Parse per-org tabs — employment + teams (no hierarchy)
    print("\n[3] Parsing per-org tabs (employment + scrum teams)...")
    org_tab_people = parse_org_roster(ORG_FILE)

    # Step 4: Parse SCRUMS.xlsx Teams Hierarchy — lead assignments
    print("\n[4] Parsing Teams Hierarchy...")
    teams_hier = parse_teams_hierarchy(SCRUMS_FILE)

    # Step 4b: Parse QA subtabs (Shefali/Rumana/Jenny) for sub-hierarchy overrides
    print("\n[4b] Parsing QA subtabs...")
    qa_subtab_overrides, qa_subtab_titles = parse_qa_subtabs(ORG_FILE)

    # Step 4c: Parse contractor locations from SCRUMS contractor list
    print("\n[4c] Parsing contractor locations...")
    contractor_locations = parse_contractor_locations(SCRUMS_FILE)
    print(f"  {len(contractor_locations)} contractor locations loaded")

    # Step 5: Build unified org datasets from on24 hierarchy + per-org enrichment
    print("\n[5] Building org datasets from on24 hierarchy...")
    org_datasets = build_from_on24(on24_people, org_tab_people, teams_hier, title_map, qa_subtab_overrides, qa_subtab_titles, contractor_locations)

    # Step 6: Build scrum index with lead identification
    print("\n[6] Building scrum team index...")
    scrum_teams = build_scrum_index(org_datasets, teams_hier)
    print(f"  {len(scrum_teams)} scrum teams found")
    for team_name in sorted(scrum_teams.keys()):
        groups = scrum_teams[team_name]
        total = sum(len(g) for g in groups.values())
        print(f"    {team_name}: {total} members")

    # Step 7: Serialize + generate HTML (named + redacted)
    print("\n[7] Generating HTML files...")
    data = make_serializable(org_datasets, scrum_teams)

    # Named version
    html_full = generate_html(data, redacted=False)
    OUTPUT_FILE.write_text(html_full, encoding='utf-8')
    print(f"  Written: {OUTPUT_FILE} ({len(html_full):,} bytes)")

    # Redacted version
    all_names = set()
    for tab_name, dataset in org_datasets.items():
        for nid, node in dataset["nodes"].items():
            if not node.get("placeholder"):
                all_names.add(node["name"])

    redacted_data = redact_data(data, all_names)
    html_redacted = generate_html(redacted_data, redacted=True)

    # Verify redaction
    leaked = verify_redaction(html_redacted, all_names)
    if leaked:
        print(f"\n  [WARN] Redaction verification: {len(leaked)} names may have leaked:")
        for name in leaked[:10]:
            print(f"    - {name}")
    else:
        print("  Redaction verification: PASSED (no names found in HTML)")

    REDACTED_FILE.write_text(html_redacted, encoding='utf-8')
    print(f"  Written: {REDACTED_FILE} ({len(html_redacted):,} bytes)")

    # Step 8: Export master data to Excel
    print("\n[8] Exporting master data to Excel...")
    export_master_excel(org_datasets, scrum_teams, teams_hier)

    # Summary
    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    for org_name in ["Product-Design", "Full QA Org", "Full Dev Org", SALESFORCE_ORG_NAME, "TPM"]:
        if org_name in org_datasets:
            ds = org_datasets[org_name]
            total = len(ds["nodes"])
            print(f"  {org_name}: {total} nodes")
    print(f"\n  Scrum teams: {len(scrum_teams)}")
    print(f"\n  Output files:")
    print(f"    {OUTPUT_FILE}")
    print(f"    {REDACTED_FILE}")
    print(f"    {MASTER_EXCEL_FILE}")
    print("\nDone!")


if __name__ == "__main__":
    main()
